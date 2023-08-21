import pandas as pd
import sqlite3
from openpyxl import load_workbook

# Chemin du fichier Excel
excel_file = "data.xlsx"

# Chemin de la base de données SQLite
sqlite_db = "Library.db"

# Connexion à la base de données
conn = sqlite3.connect(sqlite_db)

# Chargement du classeur Excel
workbook = load_workbook(excel_file)

# Récupération des noms des onglets
sheet_names = workbook.sheetnames

# Itération sur chaque onglet
for sheet_name in sheet_names:
    # Lecture des données de l'onglet en tant que DataFrame
    df = pd.read_excel(excel_file, sheet_name=sheet_name)

    # Nom de la table correspondant à l'onglet
    table_name = sheet_name.replace(" ", "_")  # Remplace les espaces par des underscores

    # Conversion du DataFrame en table SQLite
    df.to_sql(table_name, conn, index=False)

# Fermeture de la connexion à la base de données
conn.close()

print("Les tables ont été créées et les données ont été insérées avec succès.")
